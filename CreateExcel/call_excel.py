import openpyxl

fpath = r'/Users/jeong-yejin/practiceCRAWLING/CreateExcel/파일이름.xlsx'
wb = openpyxl.load_workbook(fpath)

ws = wb['오징어게임']

ws['A3'] = 456
ws['B3'] = '성기훈'

wb.save(fpath)