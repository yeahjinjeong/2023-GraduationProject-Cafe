import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet('오징어게임')

ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

wb.save(r'/Users/jeong-yejin/practiceCRAWLING/CreateExcel/파일이름.xlsx')