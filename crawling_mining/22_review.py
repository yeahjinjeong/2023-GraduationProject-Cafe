import csv
import requests
from bs4 import BeautifulSoup
from urllib.request import urlopen
import openpyxl

f = open(r'/Users/jeong-yejin/practiceCRAWLING/CreateExcel/강남구.csv',
         'r', encoding='CP949', newline='')
rdr = csv.reader(f)
# for code in rdr:
#     code[1]
# f.close()

fpath = r'/Users/jeong-yejin/practiceCRAWLING/CreateExcel/강남구.xlsx'
wb = openpyxl.Workbook(fpath)
wb.save(fpath)

for code in rdr:
    url = urlopen(
        f'https://m.place.naver.com/restaurant/{code[1]}/review/visitor?reviewItem=0')
    encoding = url.info().get_content_charset(failobj='utf-8')
    text = url.read().decode(encoding)
    wb = openpyxl.load_workbook(fpath)
    ws = wb.create_sheet(f'강남구 {code[0]}')
    soup = BeautifulSoup(text, 'html.parser')  # html번역선생님으로 수프 만듦
    links = soup.select(".zPfVt")
    for link in links:
        content = link.text
        ws.append([content])
    wb.save(fpath)
f.close()
